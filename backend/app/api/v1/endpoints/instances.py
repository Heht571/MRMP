from typing import Optional, Dict, Any, List
from uuid import UUID
from io import BytesIO
from urllib.parse import quote
from datetime import datetime
import json
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, desc
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from app.db.database import get_async_db
from app.models.instance import Instance
from app.models.meta_model_v2 import Model, ModelAttribute, OperationRecord
from app.models.import_history import ImportHistory
from app.schemas.base import (
    InstanceCreate, InstanceUpdate, InstanceResponse, InstanceDetailResponse,
    ModelResponse, AttributeResponse, PaginatedResponse
)
from app.api import deps
from app.core.config import settings
from app.lib.query_builder import QueryBuilder
from app.services.cmdb_service import ComputedAttributeService
from app.services.audit_service import AuditService
from app.services.lifecycle_service import LifecycleService
from app.models.auth import User

router = APIRouter()


async def get_model_attributes(db: AsyncSession, model_id: UUID) -> List[Dict]:
    """获取模型的属性定义"""
    result = await db.execute(
        select(ModelAttribute)
        .options(selectinload(ModelAttribute.attribute))
        .where(ModelAttribute.model_id == model_id)
        .order_by(ModelAttribute.sort_order)
    )
    model_attributes = result.scalars().all()
    
    attributes = []
    for ma in model_attributes:
        attr = ma.attribute
        attributes.append({
            "name": attr.name,
            "label": ma.override_label or attr.label,
            "type": attr.type.value if hasattr(attr.type, 'value') else attr.type,
            "is_required": ma.is_required,
            "is_unique": attr.is_unique,
            "enum_values": attr.enum_values,
            "default_value": ma.override_default or attr.default_value,
            "description": attr.description,
            "is_computed": attr.is_computed,
            "compute_expr": attr.compute_expr,
        })
    return attributes


async def validate_instance_data(
    db: AsyncSession, 
    model_id: UUID, 
    data: Dict[str, Any],
    is_update: bool = False
) -> Dict[str, Any]:
    """验证实例数据是否与元模型属性匹配"""
    attributes = await get_model_attributes(db, model_id)
    
    if not attributes:
        raise HTTPException(status_code=400, detail="该模型没有定义任何属性")
    
    valid_attr_names = set()
    required_attrs = []
    attr_definitions = {}
    
    for attr in attributes:
        attr_name = attr["name"]
        valid_attr_names.add(attr_name)
        attr_definitions[attr_name] = {
            "label": attr["label"],
            "type": attr["type"],
            "is_required": attr["is_required"],
            "enum_values": attr["enum_values"],
        }
        if attr["is_required"]:
            required_attrs.append(attr_name)
    
    if not is_update:
        missing_required = []
        for attr_name in required_attrs:
            if attr_name not in data or data[attr_name] is None or data[attr_name] == '':
                missing_required.append(attr_name)
        if missing_required:
            labels = [attr_definitions[a]["label"] for a in missing_required]
            raise HTTPException(
                status_code=400, 
                detail=f"缺少必填属性: {', '.join(labels)}"
            )
    
    invalid_attrs = set(data.keys()) - valid_attr_names
    if invalid_attrs:
        invalid_labels = list(invalid_attrs)
        raise HTTPException(
            status_code=400, 
            detail=f"以下属性未在元模型中定义: {', '.join(invalid_labels)}"
        )
    
    for attr_name, attr_def in attr_definitions.items():
        if attr_name in data and data[attr_name] is not None:
            value = data[attr_name]
            attr_type = attr_def["type"]
            
            if attr_type == "enum" and attr_def["enum_values"]:
                valid_values = [v.get("value") if isinstance(v, dict) else v for v in attr_def["enum_values"]]
                if value not in valid_values:
                    raise HTTPException(
                        status_code=400,
                        detail=f"属性 '{attr_def['label']}' 的值 '{value}' 不在有效枚举值中"
                    )
            
            elif attr_type == "number":
                try:
                    if not isinstance(value, (int, float)):
                        data[attr_name] = float(value)
                except (ValueError, TypeError):
                    raise HTTPException(
                        status_code=400,
                        detail=f"属性 '{attr_def['label']}' 必须是数字类型"
                    )
            
            elif attr_type == "boolean":
                if not isinstance(value, bool):
                    if isinstance(value, str):
                        if value.lower() in ('true', '1', 'yes', '是'):
                            data[attr_name] = True
                        elif value.lower() in ('false', '0', 'no', '否'):
                            data[attr_name] = False
                        else:
                            raise HTTPException(
                                status_code=400,
                                detail=f"属性 '{attr_def['label']}' 必须是布尔类型"
                            )
    
    return data


@router.get("/", response_model=PaginatedResponse)
async def list_instances(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    model_id: Optional[UUID] = None,
    keyword: Optional[str] = None,
    filters: Optional[str] = Query(None, description="JSON格式的过滤条件"),
    db: AsyncSession = Depends(get_async_db)
):
    query = select(Instance)
    
    if model_id:
        query = query.where(Instance.model_id == model_id)
    
    if keyword:
        keyword_filter = or_(
            Instance.name.ilike(f"%{keyword}%"),
            Instance.code.ilike(f"%{keyword}%")
        )
        query = query.where(keyword_filter)
    
    if filters:
        try:
            filter_list = json.loads(filters)
            if isinstance(filter_list, list):
                query = QueryBuilder.apply_filters(query, filter_list)
        except json.JSONDecodeError:
            pass
    
    # 优化计数查询 - 直接 count 主键，避免子查询
    count_query = select(func.count(Instance.id))
    if model_id:
        count_query = count_query.where(Instance.model_id == model_id)
    if keyword:
        keyword_filter = or_(
            Instance.name.ilike(f"%{keyword}%"),
            Instance.code.ilike(f"%{keyword}%")
        )
        count_query = count_query.where(keyword_filter)
    if filters:
        try:
            filter_list = json.loads(filters)
            if isinstance(filter_list, list):
                count_query = QueryBuilder.apply_filters(count_query, filter_list)
        except json.JSONDecodeError:
            pass

    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    query = query.offset((page - 1) * page_size).limit(page_size)
    query = query.order_by(Instance.created_at.desc())
    
    result = await db.execute(query)
    instances = result.scalars().all()
    
    return PaginatedResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=[
            InstanceResponse(
                id=i.id,
                name=i.name,
                code=i.code,
                model_id=i.model_id,
                data=i.data or {},
                status=i.status,
                version=i.version,
                created_at=i.created_at,
                updated_at=i.updated_at
            )
            for i in instances
        ]
    )


@router.post("/", response_model=InstanceResponse)
async def create_instance(
    instance_in: InstanceCreate,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(deps.get_current_active_user)
):
    """创建资源实例（数据必须与元模型属性匹配）"""
    result = await db.execute(select(Model).where(Model.id == instance_in.model_id))
    model = result.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=400, detail="指定的模型不存在")
    
    validated_data = await validate_instance_data(
        db, 
        instance_in.model_id, 
        instance_in.data or {},
        is_update=False
    )
    
    # 计算计算属性
    attributes = await get_model_attributes(db, instance_in.model_id)
    validated_data = ComputedAttributeService.calculate_attributes(validated_data, attributes)
    
    instance = Instance(
        model_id=instance_in.model_id,
        name=instance_in.name,
        code=instance_in.code,
        status=instance_in.status, # 显式设置状态
        data=validated_data
    )
    db.add(instance)
    
    # 记录审计日志
    await AuditService.log_create(db, instance, current_user.username)
    
    await db.commit()
    await db.refresh(instance)
    
    return InstanceResponse(
        id=instance.id,
        name=instance.name,
        code=instance.code,
        model_id=instance.model_id,
        data=instance.data or {},
        status=instance.status,
        version=instance.version,
        created_at=instance.created_at,
        updated_at=instance.updated_at
    )


@router.get("/{instance_id}", response_model=InstanceDetailResponse)
async def get_instance(
    instance_id: UUID,
    db: AsyncSession = Depends(get_async_db)
):
    """获取资源实例详情"""
    result = await db.execute(
        select(Instance)
        .options(selectinload(Instance.model).selectinload(Model.model_attributes).selectinload(ModelAttribute.attribute))
        .where(Instance.id == instance_id)
    )
    instance = result.scalar_one_or_none()
    if not instance:
        raise HTTPException(status_code=404, detail="实例不存在")
    
    model_response = None
    if instance.model:
        model_response = ModelResponse(
            id=instance.model.id,
            name=instance.model.name,
            code=instance.model.code,
            category=instance.model.category,
            description=instance.model.description,
            icon=instance.model.icon,
            color=instance.model.color,
            is_active=instance.model.is_active,
            is_system=getattr(instance.model, 'is_system', False),
            created_at=instance.model.created_at,
            updated_at=instance.model.updated_at,
            attributes=[
                AttributeResponse(
                    id=attr.id,
                    model_id=instance.model.id,
                    name=attr.attribute.name,
                    label=attr.override_label or attr.attribute.label,
                    type=attr.attribute.type.value if hasattr(attr.attribute.type, 'value') else attr.attribute.type,
                    is_required=attr.is_required,
                    is_unique=attr.attribute.is_unique or False,
                    is_indexed=attr.attribute.is_indexed or False,
                    is_readonly=attr.is_readonly,
                    default_value=attr.override_default or attr.attribute.default_value,
                    enum_values=attr.attribute.enum_values if attr.attribute.enum_values else None,
                    validation_regex=attr.attribute.validation_regex,
                    min_value=attr.attribute.min_value,
                    max_value=attr.attribute.max_value,
                    sort_order=attr.sort_order,
                    group_name=attr.group_name,
                    description=attr.attribute.description,
                    created_at=attr.created_at,
                    updated_at=attr.updated_at
                )
                for attr in sorted(instance.model.model_attributes, key=lambda x: x.sort_order)
            ]
        )
    
    return InstanceDetailResponse(
        id=instance.id,
        name=instance.name,
        code=instance.code,
        model_id=instance.model_id,
        data=instance.data or {},
        status=instance.status,
        version=instance.version,
        created_at=instance.created_at,
        updated_at=instance.updated_at,
        model=model_response
    )


@router.get("/{instance_id}/history")
async def get_instance_history(
    instance_id: UUID,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_async_db)
):
    """获取实例的变更历史"""
    query = select(OperationRecord).options(
        selectinload(OperationRecord.attribute_histories)
    ).where(
        OperationRecord.instance_id == instance_id
    ).order_by(desc(OperationRecord.created_at))
    
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0
    
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    records = result.scalars().all()
    
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": str(r.id),
                "operate_type": r.operate_type,
                "created_at": r.created_at,
                "created_by": r.created_by,
                "origin": r.origin,
                "changes": [
                    {
                        "attribute_name": h.attribute_name,
                        "old_value": h.old_value,
                        "new_value": h.new_value
                    } for h in r.attribute_histories
                ]
            } for r in records
        ]
    }


@router.put("/{instance_id}", response_model=InstanceResponse)
async def update_instance(
    instance_id: UUID,
    instance_in: InstanceUpdate,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(deps.get_current_active_user)
):
    """更新资源实例"""
    result = await db.execute(select(Instance).where(Instance.id == instance_id))
    instance = result.scalar_one_or_none()
    if not instance:
        raise HTTPException(status_code=404, detail="实例不存在")
    
    # 1. 乐观锁检查
    if instance_in.version is not None and instance.version != instance_in.version:
        raise HTTPException(status_code=409, detail="数据已被修改，请刷新后重试")
    
    update_data = instance_in.model_dump(exclude_unset=True)
    
    # 2. 生命周期状态检查
    if "status" in update_data and update_data["status"] != instance.status:
        LifecycleService.validate_transition(instance.status, update_data["status"])
    
    # 3. 审计日志 - 记录变更前状态
    # 注意：这里传入 update_data 用于计算 Diff
    await AuditService.log_update(db, instance, update_data, current_user.username)
    
    if "data" in update_data and update_data["data"]:
        validated_data = await validate_instance_data(
            db,
            instance.model_id,
            update_data["data"],
            is_update=True
        )
        current_data = instance.data or {}
        current_data.update(validated_data)
        
        # 重新计算计算属性
        attributes = await get_model_attributes(db, instance.model_id)
        current_data = ComputedAttributeService.calculate_attributes(current_data, attributes)
        
        update_data["data"] = current_data
    
    for field, value in update_data.items():
        if field == "version": continue # 跳过version字段的手动设置
        setattr(instance, field, value)
    
    # 自增版本号
    instance.version += 1
    
    await db.flush()
    await db.refresh(instance)
    
    return InstanceResponse(
        id=instance.id,
        name=instance.name,
        code=instance.code,
        model_id=instance.model_id,
        data=instance.data or {},
        status=instance.status,
        version=instance.version,
        created_at=instance.created_at,
        updated_at=instance.updated_at
    )


@router.delete("/{instance_id}")
async def delete_instance(
    instance_id: UUID,
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(deps.check_permissions("instance:delete"))
):
    """删除资源实例"""
    result = await db.execute(select(Instance).where(Instance.id == instance_id))
    instance = result.scalar_one_or_none()
    if not instance:
        raise HTTPException(status_code=404, detail="实例不存在")
    
    # 生命周期检查
    if not LifecycleService.can_delete(instance.status):
        raise HTTPException(status_code=400, detail=f"当前状态 '{instance.status}' 不允许删除，请先执行退网操作")
    
    # 记录审计日志
    await AuditService.log_delete(db, instance, current_user.username)
    
    await db.delete(instance)
    await db.commit()
    return {"message": "删除成功"}


@router.post("/batch-delete")
async def batch_delete_instances(
    ids: List[UUID],
    db: AsyncSession = Depends(get_async_db)
):
    """批量删除资源实例"""
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要删除的实例ID")
    
    result = await db.execute(
        select(Instance).where(Instance.id.in_(ids))
    )
    instances = result.scalars().all()
    
    if not instances:
        raise HTTPException(status_code=404, detail="未找到要删除的实例")
    
    deleted_count = 0
    error_count = 0
    
    for instance in instances:
        # 批量删除时，如果状态不允许删除，跳过并计数（或者抛出异常，这里选择跳过以支持部分成功）
        if not LifecycleService.can_delete(instance.status):
            error_count += 1
            continue
            
        await AuditService.log_delete(db, instance, "BATCH_API") # 简单记录
        await db.delete(instance)
        deleted_count += 1
    
    await db.commit()
    
    msg = f"成功删除 {deleted_count} 个实例"
    if error_count > 0:
        msg += f"，有 {error_count} 个实例因状态限制无法删除"
        
    return {"message": msg, "deleted_count": deleted_count}


@router.get("/template/{model_id}")
async def download_template(
    model_id: UUID,
    db: AsyncSession = Depends(get_async_db)
):
    """下载导入模板，包含数据校验规则"""
    result = await db.execute(select(Model).where(Model.id == model_id))
    model = result.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    attributes = await get_model_attributes(db, model_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "导入数据"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["名称*", "编码*"]
    header_comments = ["必填，实例名称", "必填，实例编码"]
    
    for attr in attributes:
        header_text = attr["label"]
        if attr["is_required"]:
            header_text += "*"
        headers.append(header_text)
        
        comment = attr["description"] or attr["label"]
        if attr["type"] == "enum" and attr["enum_values"]:
            enum_options = []
            for v in attr["enum_values"]:
                if isinstance(v, dict):
                    enum_options.append(f"{v.get('label', v.get('value'))}={v.get('value')}")
                else:
                    enum_options.append(str(v))
            comment += f"\n可选值: {', '.join(enum_options)}"
        elif attr["type"] == "boolean":
            comment += "\n可选值: 是/否"
        elif attr["type"] == "number":
            comment += "\n类型: 数字"
        elif attr["type"] == "date":
            comment += "\n格式: YYYY-MM-DD"
        header_comments.append(comment)
    
    headers.append("操作")
    header_comments.append("可选值: 删除/留空。填写'删除'将删除该记录")
    
    for col_idx, (header, comment) in enumerate(zip(headers, header_comments), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.comment = Comment(comment, "系统")
    
    example_data = ["示例名称", "EXAMPLE_001"]
    for attr in attributes:
        if attr["type"] == "string":
            example_data.append(attr["default_value"] or "示例值")
        elif attr["type"] == "number":
            example_data.append(attr["default_value"] or "0")
        elif attr["type"] == "boolean":
            example_data.append("是")
        elif attr["type"] == "enum" and attr["enum_values"]:
            first_val = attr["enum_values"][0]
            if isinstance(first_val, dict):
                example_data.append(first_val.get("label", first_val.get("value")))
            else:
                example_data.append(str(first_val))
        elif attr["type"] == "date":
            example_data.append("2024-01-01")
        else:
            example_data.append("")
    
    example_data.append("")
    
    for col_idx, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    operation_col_idx = len(headers)
    operation_col_letter = get_column_letter(operation_col_idx)
    operation_dv = DataValidation(
        type="list",
        formula1='"删除,"',
        allow_blank=True
    )
    operation_dv.error = "请选择'删除'或留空"
    operation_dv.errorTitle = "输入错误"
    operation_dv.prompt = "填写'删除'可删除该记录"
    operation_dv.promptTitle = "操作"
    operation_dv.add(f"{operation_col_letter}3:{operation_col_letter}{settings.MAX_IMPORT_ROWS}")
    ws.add_data_validation(operation_dv)
    
    for col_idx, attr in enumerate(attributes, start=3):
        col_letter = get_column_letter(col_idx)
        dv = None
        
        if attr["type"] == "enum" and attr["enum_values"]:
            enum_labels = []
            for v in attr["enum_values"]:
                if isinstance(v, dict):
                    enum_labels.append(v.get("label", v.get("value")))
                else:
                    enum_labels.append(str(v))
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(enum_labels)}"',
                allow_blank=not attr["is_required"]
            )
            dv.error = f"请从下拉列表中选择"
            dv.errorTitle = "输入错误"
            dv.prompt = f"请选择{attr['label']}"
            dv.promptTitle = attr["label"]
        elif attr["type"] == "boolean":
            dv = DataValidation(
                type="list",
                formula1='"是,否"',
                allow_blank=not attr["is_required"]
            )
            dv.error = "请输入'是'或'否'"
            dv.errorTitle = "输入错误"
        elif attr["type"] == "number":
            dv = DataValidation(
                type="decimal",
                operator="greaterThanOrEqual",
                formula1="-999999999",
                allow_blank=not attr["is_required"]
            )
            dv.error = "请输入有效的数字"
            dv.errorTitle = "输入错误"
        elif attr["type"] == "date":
            dv = DataValidation(
                type="date",
                operator="greaterThan",
                formula1="1900-01-01",
                allow_blank=not attr["is_required"]
            )
            dv.error = "请输入有效的日期格式(YYYY-MM-DD)"
            dv.errorTitle = "输入错误"
        
        if dv:
            dv.add(f"{col_letter}3:{col_letter}{settings.MAX_IMPORT_ROWS}")
            ws.add_data_validation(dv)
    
    for col_idx in range(1, len(headers) + 1):
        max_length = max(len(str(headers[col_idx - 1])), 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    
    ws.row_dimensions[1].height = 30
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{model.name}_导入模板.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/import-history/{model_id}")
async def get_import_history(
    model_id: UUID,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_async_db)
):
    """获取模型的导入历史记录"""
    result = await db.execute(select(Model).where(Model.id == model_id))
    model = result.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    count_result = await db.execute(
        select(func.count(ImportHistory.id)).where(ImportHistory.model_id == model_id)
    )
    total = count_result.scalar()
    
    result = await db.execute(
        select(ImportHistory)
        .where(ImportHistory.model_id == model_id)
        .order_by(desc(ImportHistory.started_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    histories = result.scalars().all()
    
    return {
        "items": [h.to_dict() for h in histories],
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.get("/import-history-detail/{history_id}")
async def get_import_history_detail(
    history_id: UUID,
    db: AsyncSession = Depends(get_async_db)
):
    """获取导入历史详情"""
    result = await db.execute(select(ImportHistory).where(ImportHistory.id == history_id))
    history = result.scalar_one_or_none()
    if not history:
        raise HTTPException(status_code=404, detail="导入记录不存在")
    
    return history.to_dict()


@router.post("/preview/{model_id}")
async def preview_import(
    model_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_async_db)
):
    """
    预览导入数据，返回解析后的数据和校验结果
    用于在前端展示预览，确认后再提交导入
    """
    result = await db.execute(select(Model).where(Model.id == model_id))
    model = result.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    attributes = await get_model_attributes(db, model_id)
    attr_labels = {attr["name"]: attr["label"] for attr in attributes}
    attr_map = {attr["name"]: attr for attr in attributes}
    
    content = await file.read()
    
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    col_mapping = {}
    for col_idx, header in enumerate(header_row):
        if header is None:
            continue
        header_str = str(header).strip().rstrip('*')
        if header_str == "名称":
            col_mapping["name"] = col_idx
        elif header_str == "编码":
            col_mapping["code"] = col_idx
        else:
            for attr in attributes:
                if attr["label"] == header_str:
                    col_mapping[attr["name"]] = col_idx
                    break
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    preview_data = []
    validation_errors = []
    
    for row_idx, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        
        row_data = {
            "_rowIndex": row_idx,
            "_errors": [],
            "name": None,
            "code": None,
            "data": {}
        }
        
        if "name" in col_mapping and col_mapping["name"] < len(row):
            row_data["name"] = str(row[col_mapping["name"]]).strip() if row[col_mapping["name"]] else None
        if "code" in col_mapping and col_mapping["code"] < len(row):
            row_data["code"] = str(row[col_mapping["code"]]).strip() if row[col_mapping["code"]] else None
        
        if not row_data["name"]:
            row_data["_errors"].append("名称不能为空")
        if not row_data["code"]:
            row_data["_errors"].append("编码不能为空")
        
        for attr in attributes:
            attr_name = attr["name"]
            if attr_name in col_mapping and col_mapping[attr_name] < len(row):
                value = row[col_mapping[attr_name]]
                if value is not None and str(value).strip() != '':
                    attr_type = attr["type"]
                    if attr_type == "number":
                        try:
                            value = float(value) if '.' in str(value) else int(value)
                        except:
                            row_data["_errors"].append(f"{attr['label']}必须是数字")
                    elif attr_type == "boolean":
                        if isinstance(value, str):
                            value = value.lower() in ('true', '1', 'yes', '是')
                        else:
                            value = bool(value)
                    elif attr_type == "enum" and attr["enum_values"]:
                        matched = False
                        for v in attr["enum_values"]:
                            if isinstance(v, dict):
                                if v.get("label") == value or v.get("value") == value:
                                    value = v.get("value")
                                    matched = True
                                    break
                            elif str(v) == str(value):
                                matched = True
                                break
                        if not matched and value:
                            row_data["_errors"].append(f"{attr['label']}的值不在可选范围内")
                    row_data["data"][attr_name] = value
                elif attr["is_required"]:
                    row_data["_errors"].append(f"{attr['label']}是必填项")
        
        if row_data["_errors"]:
            validation_errors.extend([f"第{row_idx}行: {err}" for err in row_data["_errors"]])
        
        preview_data.append(row_data)
    
    existing_codes = set()
    if preview_data:
        codes = [r["code"] for r in preview_data if r["code"]]
        if codes:
            result = await db.execute(
                select(Instance.code).where(
                    Instance.model_id == model_id,
                    Instance.code.in_(codes)
                )
            )
            existing_codes = set(r[0] for r in result.fetchall())
    
    for row_data in preview_data:
        if row_data["code"] in existing_codes:
            row_data["_status"] = "update"
        else:
            row_data["_status"] = "create"
    
    return {
        "total": len(preview_data),
        "createCount": sum(1 for r in preview_data if r["_status"] == "create"),
        "updateCount": sum(1 for r in preview_data if r["_status"] == "update"),
        "errorCount": len(validation_errors),
        "errors": validation_errors[:20],
        "data": preview_data[:100],
        "columns": ["名称", "编码"] + [attr["label"] for attr in attributes],
        "fieldMapping": col_mapping
    }


@router.post("/import/{model_id}")
async def import_instances(
    model_id: UUID,
    file: UploadFile = File(...),
    mode: str = Query("upsert", description="导入模式: create_only(仅新增), update_only(仅更新), upsert(新增或更新)"),
    db: AsyncSession = Depends(get_async_db)
):
    """
    导入Excel数据
    
    导入模式:
    - create_only: 仅新增，遇到重复记录则跳过并记录错误
    - update_only: 仅更新，根据编码匹配现有记录进行更新
    - upsert: 智能处理，编码存在则更新，不存在则新增
    """
    result = await db.execute(select(Model).where(Model.id == model_id))
    model = result.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    attributes = await get_model_attributes(db, model_id)
    attr_map = {attr["name"]: attr for attr in attributes}
    unique_attrs = [attr["name"] for attr in attributes if attr.get("is_unique")]
    
    content = await file.read()
    
    import_history = ImportHistory(
        model_id=model_id,
        file_name=file.filename,
        file_size=len(content),
        import_mode=mode,
        status="processing"
    )
    db.add(import_history)
    await db.commit()
    await db.refresh(import_history)
    
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    success_count = 0
    update_count = 0
    skip_count = 0
    delete_count = 0
    error_list = []
    
    for row_idx, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row[:2]):
            continue
        
        try:
            name = str(row[0]).strip() if row[0] else None
            code = str(row[1]).strip() if row[1] else None
            operation = str(row[-1]).strip().lower() if len(row) > len(attributes) + 2 and row[-1] else None
            
            if not name:
                error_list.append(f"第{row_idx}行: 名称不能为空")
                continue
            if not code:
                error_list.append(f"第{row_idx}行: 编码不能为空")
                continue
            
            if operation == "删除":
                existing_instance = await db.execute(
                    select(Instance).where(
                        Instance.model_id == model_id,
                        Instance.code == code
                    )
                )
                to_delete = existing_instance.scalar_one_or_none()
                if to_delete:
                    await db.delete(to_delete)
                    delete_count += 1
                    success_count += 1
                else:
                    error_list.append(f"第{row_idx}行: 编码 '{code}' 不存在，无法删除")
                    skip_count += 1
                continue
            
            data = {}
            for col_idx, attr in enumerate(attributes, start=2):
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None and str(value).strip() != '':
                        attr_type = attr["type"]
                        if attr_type == "number":
                            try:
                                value = float(value) if '.' in str(value) else int(value)
                            except:
                                pass
                        elif attr_type == "boolean":
                            if isinstance(value, str):
                                value = value.lower() in ('true', '1', 'yes', '是')
                            else:
                                value = bool(value)
                        elif attr_type == "enum" and attr["enum_values"]:
                            for v in attr["enum_values"]:
                                if isinstance(v, dict):
                                    if v.get("label") == value or v.get("value") == value:
                                        value = v.get("value")
                                        break
                                elif str(v) == str(value):
                                    break
                        data[attr["name"]] = value
            
            existing_instance = await db.execute(
                select(Instance).where(
                    Instance.model_id == model_id,
                    Instance.code == code
                )
            )
            existing = existing_instance.scalar_one_or_none()
            
            for unique_attr in unique_attrs:
                if unique_attr in data and data[unique_attr] is not None:
                    unique_check = await db.execute(
                        select(Instance).where(
                            Instance.model_id == model_id,
                            Instance.data[unique_attr].astext == str(data[unique_attr])
                        )
                    )
                    unique_existing = unique_check.scalar_one_or_none()
                    if unique_existing and (not existing or unique_existing.id != existing.id):
                        error_list.append(f"第{row_idx}行: 唯一属性 '{unique_attr}' 的值 '{data[unique_attr]}' 已存在")
                        skip_count += 1
                        continue
            
            if existing:
                if mode == "create_only":
                    error_list.append(f"第{row_idx}行: 编码 '{code}' 已存在，跳过")
                    skip_count += 1
                    continue
                
                validated_data = await validate_instance_data(db, model_id, data, is_update=True)
                existing.name = name
                existing.data = {**existing.data, **validated_data}
                update_count += 1
                success_count += 1
            else:
                if mode == "update_only":
                    error_list.append(f"第{row_idx}行: 编码 '{code}' 不存在，跳过")
                    skip_count += 1
                    continue
                
                validated_data = await validate_instance_data(db, model_id, data, is_update=False)
                instance = Instance(
                    model_id=model_id,
                    name=name,
                    code=code,
                    data=validated_data
                )
                db.add(instance)
                success_count += 1
            
        except HTTPException as e:
            error_list.append(f"第{row_idx}行: {e.detail}")
        except Exception as e:
            error_list.append(f"第{row_idx}行: {str(e)}")
    
    await db.commit()
    
    import_history.total_rows = len(rows)
    import_history.created_count = success_count - update_count - delete_count
    import_history.updated_count = update_count
    import_history.deleted_count = delete_count
    import_history.skipped_count = skip_count
    import_history.error_count = len(error_list)
    import_history.errors = error_list[:100]
    import_history.status = "completed"
    import_history.finished_at = datetime.utcnow()
    await db.commit()
    
    return {
        "id": str(import_history.id),
        "success": success_count,
        "created": success_count - update_count - delete_count,
        "updated": update_count,
        "deleted": delete_count,
        "skipped": skip_count,
        "errors": error_list,
        "total": len(rows)
    }


@router.get("/export/{model_id}")
async def export_instances(
    model_id: UUID,
    keyword: Optional[str] = Query(None, description="搜索关键字"),
    ids: Optional[str] = Query(None, description="指定导出的实例ID，逗号分隔"),
    db: AsyncSession = Depends(get_async_db)
):
    """导出Excel数据，支持关键字筛选和指定ID导出"""
    result = await db.execute(select(Model).where(Model.id == model_id))
    model = result.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    attributes = await get_model_attributes(db, model_id)
    
    query = select(Instance).where(Instance.model_id == model_id)
    
    if ids:
        id_list = [UUID(id.strip()) for id in ids.split(",") if id.strip()]
        query = query.where(Instance.id.in_(id_list))
    elif keyword:
        query = query.where(
            or_(
                Instance.name.ilike(f"%{keyword}%"),
                Instance.code.ilike(f"%{keyword}%")
            )
        )
    
    query = query.order_by(Instance.created_at.desc())
    result = await db.execute(query)
    instances = result.scalars().all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "导出数据"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["名称", "编码"]
    for attr in attributes:
        headers.append(attr["label"])
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, instance in enumerate(instances, start=2):
        ws.cell(row=row_idx, column=1, value=instance.name).border = thin_border
        ws.cell(row=row_idx, column=2, value=instance.code).border = thin_border
        
        for col_idx, attr in enumerate(attributes, start=3):
            value = instance.data.get(attr["name"], "") if instance.data else ""
            
            if attr["type"] == "enum" and attr["enum_values"] and value:
                for v in attr["enum_values"]:
                    if isinstance(v, dict) and v.get("value") == value:
                        value = v.get("label", value)
                        break
            elif attr["type"] == "boolean":
                value = "是" if value else "否"
            
            ws.cell(row=row_idx, column=col_idx, value=value).border = thin_border
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{model.name}_导出数据.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
